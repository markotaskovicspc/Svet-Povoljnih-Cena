from __future__ import annotations

import json
import math
import re
import subprocess
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "outputs" / "Book16_export.xlsx"
OUT_DIR = ROOT / "outputs"
PRODUCT_PHOTOS_DIR = ROOT / "Product photos"
WORKBOOK_OUT = OUT_DIR / "Svet_akcija_product_inventory_and_UX_plan.xlsx"
PRODUCTS_JSON_OUT = OUT_DIR / "svet-akcija-products.json"
CATEGORIES_JSON_OUT = OUT_DIR / "svet-akcija-categories.json"
GROUPS_JSON_OUT = OUT_DIR / "svet-akcija-groups.json"
CHECKS_JSON_OUT = OUT_DIR / "svet-akcija-data-checks.json"
PRODUCT_ASSETS_JSON_OUT = OUT_DIR / "svet-akcija-product-assets.json"
MEDIA_UPLOAD_MANIFEST_JSON_OUT = OUT_DIR / "svet-akcija-product-media-upload-manifest.json"
SUPABASE_SQL_OUT = OUT_DIR / "Svet_akcija_supabase_import.sql"
IMPORT_BATCH = "book16_2026_05"
PROMO_ACTION_ID = "action-svet-akcija-maj-2026"
PROMO_ACTION_SLUG = "akcija"
PROMO_ACTION_NAME = "Svet akcija - maj 2026"
PROMO_ACTION_START = "2026-05-01 00:00:00"
PROMO_ACTION_END = "2026-05-31 23:59:59"


REQUIRED_CHECK_FIELDS = [
    "Šifra",
    "Kategorija",
    "Grupa",
    "Kratki naziv",
    "Opis",
    "Kolekcija (brend)",
    "Boja 1",
    "Boja 2",
    "Bar kod",
    "MPC redovna",
    "Akcijska MPC",
    "Važenje akcijske cene od",
    "Važenje akcijske cene do",
]

PLACEHOLDERS = {"9", "0", "/", "-", "N/A", "n/a", "NA", "na"}
SUSPICIOUS_FIELDS = [
    "Kolekcija (brend)",
    "Atribut 1",
    "Atribut 2",
    "Boja 1",
    "Boja 2",
    "DC (lager)",
]
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".avif"}
MEDIA_VARIANTS = {
    "thumbUrl": ("thumb", 160),
    "cardUrl": ("card", 640),
    "pdpUrl": ("pdp", 1280),
}
LFS_POINTER_RE = re.compile(r"^(version https://git-lfs\.github\.com/spec/v1\b|oid sha256:)", re.IGNORECASE)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value == "")


def to_display(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def json_dump(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def compact_list(values: list[Any], limit: int = 80) -> str:
    cleaned = [str(v) for v in values]
    if len(cleaned) > limit:
        return ", ".join(cleaned[:limit]) + f" ... (+{len(cleaned) - limit} more)"
    return ", ".join(cleaned)


def safe_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def safe_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).date()
        except ValueError:
            continue
    return None


def campaign_status(starts_at: Any, ends_at: Any, today: date | None = None) -> str:
    today = today or date.today()
    start = safe_date(starts_at)
    end = safe_date(ends_at)
    if start is None or end is None:
        return "invalid"
    if end < today:
        return "expired"
    if start > today:
        return "future"
    return "current"


def is_lfs_pointer_text(value: str | None) -> bool:
    return bool(value and LFS_POINTER_RE.search(value.strip()))


def normalize_sku(value: Any) -> str:
    return str(value).strip().replace(".", "")


def natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value)]


def image_sort_key(path: Path) -> tuple[int, list[Any]]:
    match = re.match(r"^\s*(\d+)(?:$|[\s._-])", path.stem)
    if match:
        return (0, [int(match.group(1)), *natural_key(path.name)])
    return (1, natural_key(path.name))


def slugify_storage_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_text).strip("-").lower()
    return slug or "image"


def variant_storage_paths(storage_path: str) -> dict[str, str]:
    source = Path(storage_path)
    source_base = str(source.with_suffix("")).replace("\\", "/")
    return {
        column: f"variants/{variant}/{source_base}-{width}.webp"
        for column, (variant, width) in MEDIA_VARIANTS.items()
    }


def clean_docx_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u00a0", " ")
    lines = []
    blank = False
    for raw_line in text.split("\n"):
        line = re.sub(r"[ \t]+", " ", raw_line).strip()
        if not line:
            if lines and not blank:
                lines.append("")
            blank = True
            continue
        lines.append(line)
        blank = False
    return "\n".join(lines).strip()


def extract_docx_text(path: Path) -> str:
    completed = subprocess.run(
        ["textutil", "-convert", "txt", "-stdout", str(path)],
        check=True,
        capture_output=True,
        text=True,
    )
    return clean_docx_text(completed.stdout)


def as_source_record(row: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    return {h: to_display(row.get(h)) for h in headers}


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], widths: dict[int, int] | None = None) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="173E43")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = widths or {}
    for col_idx, _ in enumerate(headers, 1):
        width = widths.get(col_idx)
        if width is None:
            values = [headers[col_idx - 1]] + [r[col_idx - 1] for r in rows[:100] if col_idx <= len(r)]
            longest = max(len(str(v)) if v is not None else 0 for v in values)
            width = max(12, min(42, longest + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_product_assets(product_skus: set[str]) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    assets_by_sku: dict[str, dict[str, Any]] = {}
    unmatched_folders = []
    duplicate_folders = []
    extraction_failures = []

    if not PRODUCT_PHOTOS_DIR.exists():
        return {}, {
            "assetFolderCount": 0,
            "matchedAssetSkus": 0,
            "missingAssetSkus": sorted(product_skus, key=natural_key),
            "unmatchedAssetFolders": [{"folder": str(PRODUCT_PHOTOS_DIR), "normalizedSku": None}],
            "duplicateAssetFolders": [],
            "docxExtractionFailures": [],
        }

    folders = sorted([p for p in PRODUCT_PHOTOS_DIR.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name))
    for folder in folders:
        sku = normalize_sku(folder.name)
        warnings = []
        if sku not in product_skus:
            unmatched_folders.append({"folder": folder.name, "normalizedSku": sku})
            continue
        if sku in assets_by_sku:
            duplicate_folders.append({"sku": sku, "folder": folder.name, "previousFolder": assets_by_sku[sku]["sourceFolder"]})
            continue

        files = sorted([p for p in folder.iterdir() if p.is_file()], key=lambda p: natural_key(p.name))
        docs = [p for p in files if p.suffix.lower() == ".docx"]
        images = sorted([p for p in files if p.suffix.lower() in IMAGE_EXTENSIONS], key=image_sort_key)
        other_files = [p.name for p in files if p.suffix.lower() not in IMAGE_EXTENSIONS and p.suffix.lower() != ".docx"]

        docx_file = docs[0] if docs else None
        long_description = ""
        if not docs:
            warnings.append("missing:docx")
        elif len(docs) > 1:
            warnings.append("multiple:docx")

        if docx_file is not None:
            try:
                long_description = extract_docx_text(docx_file)
                if not long_description:
                    warnings.append("empty:docx_text")
                elif is_lfs_pointer_text(long_description):
                    warnings.append("broken:lfs_pointer_description")
                    long_description = ""
            except (subprocess.CalledProcessError, OSError) as exc:
                warnings.append("failed:docx_extraction")
                extraction_failures.append({"sku": sku, "folder": folder.name, "docxFile": docx_file.name, "error": str(exc)})

        if not images:
            warnings.append("missing:images")

        media = []
        for index, image_path in enumerate(images):
            order_label = f"{index + 1:03d}"
            storage_path = f"products/{sku}/{order_label}-{slugify_storage_name(image_path.stem)}{image_path.suffix.lower()}"
            media.append(
                {
                    "id": f"sa-media-{sku}-{order_label}",
                    "order": index,
                    "kind": "IMAGE",
                    "sourceFile": image_path.name,
                    "localPath": str(image_path.relative_to(ROOT)),
                    "absoluteLocalPath": str(image_path),
                    "storagePath": storage_path,
                    **variant_storage_paths(storage_path),
                }
            )

        assets_by_sku[sku] = {
            "sku": sku,
            "sourceFolder": folder.name,
            "normalizedFolderSku": sku,
            "docxFile": docx_file.name if docx_file else None,
            "longDescription": long_description,
            "descriptionCharCount": len(long_description),
            "media": media,
            "ignoredFiles": other_files,
            "warnings": warnings,
        }

    missing_asset_skus = sorted(product_skus.difference(assets_by_sku), key=natural_key)
    summary = {
        "assetFolderCount": len(folders),
        "matchedAssetSkus": len(assets_by_sku),
        "missingAssetSkus": missing_asset_skus,
        "unmatchedAssetFolders": unmatched_folders,
        "duplicateAssetFolders": duplicate_folders,
        "docxExtractionFailures": extraction_failures,
    }
    return assets_by_sku, summary


def sql_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def render_supabase_import_sql(
    products: list[dict[str, Any]],
    headers: list[str],
    asset_by_sku: dict[str, dict[str, Any]],
    duplicate_barcode_values: set[str],
) -> str:
    source_rows = []
    for r in products:
        flags = []
        for field in REQUIRED_CHECK_FIELDS:
            if is_blank(r.get(field)):
                flags.append(f"missing:{field}")
        if not is_blank(r.get("Bar kod")) and str(r.get("Bar kod")) in duplicate_barcode_values:
            flags.append("duplicate:Bar kod")
        regular = safe_number(r.get("MPC redovna"))
        sale = safe_number(r.get("Akcijska MPC"))
        if regular is not None and sale is not None:
            if sale > regular:
                flags.append("price:action_higher_than_regular")
            if sale == regular:
                flags.append("price:regular_equals_action")
            if sale < regular:
                status = campaign_status(
                    r.get("Važenje akcijske cene od"),
                    r.get("Važenje akcijske cene do"),
                )
                if status != "current":
                    flags.append(f"campaign:{status}")
        for field in headers:
            value = r.get(field)
            if value is not None and str(value).strip() in PLACEHOLDERS:
                flags.append(f"suspicious_placeholder:{field}")
        source_rows.append({"source_row": r["_source_row"], "flags": flags, **as_source_record(r, headers)})

    asset_rows = [
        {
            "sku": sku,
            "longDescription": asset["longDescription"],
            "mediaCount": len(asset["media"]),
        }
        for sku, asset in sorted(asset_by_sku.items(), key=lambda item: natural_key(item[0]))
    ]
    media_rows = []
    for sku, asset in sorted(asset_by_sku.items(), key=lambda item: natural_key(item[0])):
        product_title = next((to_display(p.get("Kratki naziv")) for p in products if normalize_sku(p.get("Šifra")) == sku), sku)
        for media in asset["media"]:
            media_rows.append(
                {
                    "sku": sku,
                    "media_id": media["id"],
                    "media_order": media["order"],
                    "url": media["storagePath"],
                    "thumb_url": media["thumbUrl"],
                    "card_url": media["cardUrl"],
                    "pdp_url": media["pdpUrl"],
                    "alt": f"{product_title} - slika {media['order'] + 1}",
                }
            )

    source_json = json.dumps(source_rows, ensure_ascii=False, separators=(",", ":"))
    asset_json = json.dumps(asset_rows, ensure_ascii=False, separators=(",", ":"))
    media_json = json.dumps(media_rows, ensure_ascii=False, separators=(",", ":"))
    source_columns = ",\n    ".join([f"{sql_ident(h)} TEXT" for h in headers])
    insert_columns = ", ".join(sql_ident(h) for h in headers)
    select_columns = ", ".join(sql_ident(h) for h in headers)
    update_columns = ",\n    ".join([f"{sql_ident(h)} = EXCLUDED.{sql_ident(h)}" for h in headers])

    return f"""-- Svet Akcija Book16 catalog import for Supabase
-- Generated from /Users/luka/Documents/Book16.numbers plus Product photos/ DOCX and images.
-- Paste this whole file into Supabase SQL Editor and run it once.
-- It is idempotent: re-running updates the same SKUs and replaces only generated ProductMedia rows.
--
-- ProductMedia.url stores Supabase Storage object paths such as products/1081/001-1.png.
-- Upload local files according to outputs/svet-akcija-product-media-upload-manifest.json.
--
-- Launch-safety rules:
-- 1) Stock comes only from owner source data in "DC (lager)" when it is a non-negative integer.
--    Missing/invalid stock imports as stock=0 and isActive=false.
-- 2) Products without generated media import as inactive until assets are provided.
-- 3) Expired campaign prices are not mapped to Product.salePrice or Product.actionId.
--
-- Important preservation rule:
-- 1) public.svet_akcija_product_import stores every original source column exactly as text.
-- 2) The mapped Product insert is a website layer. Because Product.barcode is UNIQUE and the
--    workbook contains duplicate barcode 6979022547901 for SKUs 1108 and 1026, duplicate
--    barcodes are stored exactly in the raw table but left NULL in Product.barcode.

BEGIN;

CREATE EXTENSION IF NOT EXISTS pg_trgm;

CREATE OR REPLACE FUNCTION public.svet_akcija_slugify(value TEXT)
RETURNS TEXT
LANGUAGE SQL
IMMUTABLE
AS $$
  SELECT trim(both '-' from regexp_replace(
    lower(translate(coalesce(value, ''), 'ŠĐČĆŽšđčćž', 'SDCCZsdccz')),
    '[^a-z0-9]+', '-', 'g'
  ));
$$;

CREATE OR REPLACE FUNCTION public.svet_akcija_category_slug(value TEXT)
RETURNS TEXT
LANGUAGE SQL
IMMUTABLE
AS $$
  SELECT CASE
    WHEN value = 'Kućni aparati i beauty oprema' THEN 'kucni-aparati'
    ELSE public.svet_akcija_slugify(value)
  END;
$$;

CREATE OR REPLACE FUNCTION public.svet_akcija_stable_id(prefix TEXT, value TEXT)
RETURNS TEXT
LANGUAGE SQL
IMMUTABLE
AS $$
  SELECT prefix || '-' || left(md5(coalesce(value, '')), 24);
$$;

-- Defensive compatibility for projects that have not yet applied the Excel-field migration.
ALTER TABLE public."Product" ADD COLUMN IF NOT EXISTS barcode TEXT;
ALTER TABLE public."Product" ADD COLUMN IF NOT EXISTS "sizeLabel" TEXT;
ALTER TABLE public."Product" ADD COLUMN IF NOT EXISTS "colorPrimary" TEXT;
ALTER TABLE public."Product" ADD COLUMN IF NOT EXISTS "colorSecondary" TEXT;
CREATE UNIQUE INDEX IF NOT EXISTS "Product_barcode_key" ON public."Product"(barcode);

CREATE TABLE IF NOT EXISTS public.svet_akcija_product_import (
  import_batch TEXT NOT NULL DEFAULT '{IMPORT_BATCH}',
  source_file TEXT NOT NULL DEFAULT 'Book16.numbers',
  source_row INT NOT NULL,
  flags JSONB NOT NULL DEFAULT '[]'::jsonb,
  "Šifra" TEXT,
  "Kategorija" TEXT,
  "Grupa" TEXT,
  "Dobavljač" TEXT,
  "Kolekcija (brend)" TEXT,
  "Opis" TEXT,
  "Kratki naziv" TEXT,
  "Atribut 1" TEXT,
  "Atribut 2" TEXT,
  "Boja 1" TEXT,
  "Boja 2" TEXT,
  "DC (lager)" TEXT,
  "Bar kod" TEXT,
  "MPC redovna" TEXT,
  "Akcijska MPC" TEXT,
  "Važenje akcijske cene od" TEXT,
  "Važenje akcijske cene do" TEXT,
  imported_at TIMESTAMP(3) NOT NULL DEFAULT now(),
  PRIMARY KEY (import_batch, "Šifra")
);

WITH source_rows AS (
  SELECT *
  FROM jsonb_to_recordset($catalog_json${source_json}$catalog_json$::jsonb) AS x(
    source_row INT,
    flags JSONB,
    {source_columns}
  )
)
INSERT INTO public.svet_akcija_product_import (source_row, flags, {insert_columns})
SELECT source_row, flags, {select_columns}
FROM source_rows
ON CONFLICT (import_batch, "Šifra") DO UPDATE SET
    {update_columns},
    source_row = EXCLUDED.source_row,
    flags = EXCLUDED.flags,
    imported_at = now();

-- Website category mapping. Public navigation uses the current site URLs, while the exact
-- workbook category value remains preserved in public.svet_akcija_product_import."Kategorija".
WITH roots(label, slug, sort_order) AS (
  VALUES
  ('Nameštaj', 'namestaj', 1),
  ('Sve za kuću', 'sve-za-kucu', 2),
  ('Kućni aparati', 'kucni-aparati', 3),
  ('Moda i putovanja', 'moda-i-putovanja', 4)
)
INSERT INTO public."Category" (id, slug, name, "parentId", "order", path, level, "createdAt", "updatedAt")
SELECT
  'cat-' || slug,
  slug,
  label,
  NULL,
  sort_order,
  '/' || slug,
  0,
  now(),
  now()
FROM roots
ON CONFLICT (slug) DO UPDATE SET
  name = EXCLUDED.name,
  "parentId" = EXCLUDED."parentId",
  "order" = EXCLUDED."order",
  path = EXCLUDED.path,
  level = EXCLUDED.level,
  "updatedAt" = now();

WITH children(root_label, root_slug, label, slug, sort_order) AS (
  VALUES
  ('Nameštaj', 'namestaj', 'Baštenski nameštaj', 'bastenski-namestaj', 1),
  ('Nameštaj', 'namestaj', 'Kancelarija', 'kancelarija', 2),
  ('Nameštaj', 'namestaj', 'Trpezarija', 'trpezarija', 3),
  ('Nameštaj', 'namestaj', 'Dnevna soba', 'dnevna-soba', 4),
  ('Nameštaj', 'namestaj', 'Predsoblje', 'predsoblje', 5),
  ('Nameštaj', 'namestaj', 'Gejming', 'gejming', 6),
  ('Nameštaj', 'namestaj', 'Spavaća soba', 'spavaca-soba', 7),
  ('Sve za kuću', 'sve-za-kucu', 'Bazeni', 'bazeni', 8),
  ('Sve za kuću', 'sve-za-kucu', 'Alat', 'alat', 9),
  ('Sve za kuću', 'sve-za-kucu', 'Rasveta', 'rasveta', 10),
  ('Sve za kuću', 'sve-za-kucu', 'Čišćenje i održavanje', 'ciscenje-i-odrzavanje', 11),
  ('Sve za kuću', 'sve-za-kucu', 'Dekoracija', 'dekoracija', 12),
  ('Sve za kuću', 'sve-za-kucu', 'Kupatilo', 'kupatilo', 13),
  ('Sve za kuću', 'sve-za-kucu', 'Tepisi', 'tepisi', 14),
  ('Kućni aparati', 'kucni-aparati', 'Kafe aparati', 'kafe-aparati', 15),
  ('Kućni aparati', 'kucni-aparati', 'Lepota i nega', 'lepota-i-nega', 16),
  ('Kućni aparati', 'kucni-aparati', 'Hlađenje i grejanje', 'hladjenje-i-grejanje', 17),
  ('Kućni aparati', 'kucni-aparati', 'Priprema hrane', 'priprema-hrane', 18),
  ('Kućni aparati', 'kucni-aparati', 'Kuvanje i pečenje', 'kuvanje-i-pecenje', 19),
  ('Kućni aparati', 'kucni-aparati', 'Pegle', 'pegle', 20),
  ('Kućni aparati', 'kucni-aparati', 'Usisivači', 'usisivaci', 21),
  ('Kućni aparati', 'kucni-aparati', 'Prečišćivači vazduha', 'preciscivaci-vazduha', 22),
  ('Kućni aparati', 'kucni-aparati', 'Aparati za vodu', 'aparati-za-vodu', 23),
  ('Moda i putovanja', 'moda-i-putovanja', 'Ženske torbe', 'zenske-torbe', 24),
  ('Moda i putovanja', 'moda-i-putovanja', 'Ženske čarape', 'zenske-carape', 25),
  ('Moda i putovanja', 'moda-i-putovanja', 'Aksesoari', 'aksesoari', 26),
  ('Moda i putovanja', 'moda-i-putovanja', 'Koferi', 'koferi', 27)
)
INSERT INTO public."Category" (id, slug, name, "parentId", "order", path, level, "createdAt", "updatedAt")
SELECT
  'cat-' || root_slug || '-' || slug,
  root_slug || '-' || slug,
  root_label || ' / ' || label,
  (SELECT c.id FROM public."Category" c WHERE c.slug = children.root_slug LIMIT 1),
  sort_order,
  '/' || root_slug || '/' || slug,
  1,
  now(),
  now()
FROM children
ON CONFLICT (slug) DO UPDATE SET
  name = EXCLUDED.name,
  "parentId" = EXCLUDED."parentId",
  "order" = EXCLUDED."order",
  path = EXCLUDED.path,
  level = EXCLUDED.level,
  "updatedAt" = now();

-- Groups, collections/brands, and suppliers from exact source values.
INSERT INTO public."Group" (id, slug, name)
SELECT DISTINCT
  public.svet_akcija_stable_id('grp', "Grupa"),
  public.svet_akcija_slugify("Grupa"),
  "Grupa"
FROM public.svet_akcija_product_import
WHERE import_batch = '{IMPORT_BATCH}' AND nullif("Grupa", '') IS NOT NULL
ON CONFLICT (slug) DO UPDATE SET name = EXCLUDED.name;

INSERT INTO public."Collection" (id, slug, name)
SELECT DISTINCT
  public.svet_akcija_stable_id('col', "Kolekcija (brend)"),
  public.svet_akcija_slugify("Kolekcija (brend)"),
  "Kolekcija (brend)"
FROM public.svet_akcija_product_import
WHERE import_batch = '{IMPORT_BATCH}'
  AND nullif("Kolekcija (brend)", '') IS NOT NULL
  AND trim("Kolekcija (brend)") NOT IN ('9','0','/','-','N/A','n/a','NA','na')
ON CONFLICT (slug) DO UPDATE SET name = EXCLUDED.name;

INSERT INTO public."Supplier" (id, name, enabled, "createdAt", "updatedAt")
SELECT DISTINCT
  public.svet_akcija_stable_id('sup', "Dobavljač"),
  "Dobavljač",
  true,
  now(),
  now()
FROM public.svet_akcija_product_import
WHERE import_batch = '{IMPORT_BATCH}' AND nullif("Dobavljač", '') IS NOT NULL
ON CONFLICT (name) DO UPDATE SET
  name = EXCLUDED.name,
  enabled = true,
  "updatedAt" = now();

INSERT INTO public."Action" (id, slug, name, kind, "startsAt", "endsAt", "isHero", "sortOrder", "createdAt", "updatedAt")
SELECT
  '{PROMO_ACTION_ID}',
  '{PROMO_ACTION_SLUG}',
  '{PROMO_ACTION_NAME}',
  'AKCIJA',
  TIMESTAMP '{PROMO_ACTION_START}',
  TIMESTAMP '{PROMO_ACTION_END}',
  false,
  1,
  now(),
  now()
WHERE TIMESTAMP '{PROMO_ACTION_START}' <= CURRENT_TIMESTAMP
  AND TIMESTAMP '{PROMO_ACTION_END}' >= CURRENT_TIMESTAMP
ON CONFLICT (slug) DO UPDATE SET
  name = EXCLUDED.name,
  kind = EXCLUDED.kind,
  "startsAt" = EXCLUDED."startsAt",
  "endsAt" = EXCLUDED."endsAt",
  "updatedAt" = now();

-- Product storefront mapping. Exact source fields remain in the raw import table above.
WITH product_assets AS (
  SELECT *
  FROM jsonb_to_recordset($asset_json${asset_json}$asset_json$::jsonb) AS x(
    sku TEXT,
    "longDescription" TEXT,
    "mediaCount" INT
  )
), raw AS (
  SELECT
    r.*,
    count(*) OVER (PARTITION BY nullif(r."Bar kod", '')) AS barcode_count
  FROM public.svet_akcija_product_import r
  WHERE r.import_batch = '{IMPORT_BATCH}'
), mapped AS (
  SELECT
    'sa-prod-' || raw."Šifra" AS id,
    raw."Šifra" AS sku,
    CASE
      WHEN nullif(raw."Bar kod", '') IS NULL THEN NULL
      WHEN barcode_count > 1 THEN NULL
      WHEN EXISTS (
        SELECT 1 FROM public."Product" p
        WHERE p.barcode = raw."Bar kod" AND p.sku <> raw."Šifra"
      ) THEN NULL
      ELSE raw."Bar kod"
    END AS barcode,
    public.svet_akcija_slugify(raw."Kratki naziv") || '-' || public.svet_akcija_slugify(raw."Šifra") AS slug,
    raw."Kratki naziv" AS name,
    COALESCE(NULLIF(product_assets."longDescription", ''), raw."Opis") AS description,
    raw."Opis" AS short_description,
    NULLIF(raw."Boja 1", '') AS color_primary,
    NULLIF(raw."Boja 2", '') AS color_secondary,
    (SELECT g.id FROM public."Group" g WHERE g.slug = public.svet_akcija_slugify(raw."Grupa") LIMIT 1) AS group_id,
    (SELECT c.id FROM public."Collection" c WHERE c.slug = public.svet_akcija_slugify(raw."Kolekcija (brend)") LIMIT 1) AS collection_id,
    NULLIF(raw."MPC redovna", '')::numeric(12,2) AS full_price,
    CASE
      WHEN NULLIF(raw."Akcijska MPC", '') IS NOT NULL
       AND NULLIF(raw."Akcijska MPC", '')::numeric(12,2) < NULLIF(raw."MPC redovna", '')::numeric(12,2)
       AND raw."Važenje akcijske cene od" ~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}'
       AND raw."Važenje akcijske cene do" ~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}'
       AND raw."Važenje akcijske cene od"::timestamp <= CURRENT_TIMESTAMP
       AND raw."Važenje akcijske cene do"::timestamp >= CURRENT_TIMESTAMP
      THEN NULLIF(raw."Akcijska MPC", '')::numeric(12,2)
      ELSE NULL
    END AS sale_price,
    CASE WHEN nullif(raw."DC (lager)", '') ~ '^[0-9]+$' THEN raw."DC (lager)"::int ELSE 0 END AS stock,
    CASE WHEN nullif(raw."DC (lager)", '') ~ '^[0-9]+$' THEN raw."DC (lager)"::int ELSE NULL END AS supplier_stock,
    COALESCE(product_assets."mediaCount", 0) AS media_count,
    (SELECT s.id FROM public."Supplier" s WHERE s.name = raw."Dobavljač" LIMIT 1) AS supplier_id,
    CASE
      WHEN NULLIF(raw."Akcijska MPC", '') IS NOT NULL
       AND NULLIF(raw."Akcijska MPC", '')::numeric(12,2) < NULLIF(raw."MPC redovna", '')::numeric(12,2)
       AND raw."Važenje akcijske cene od" ~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}'
       AND raw."Važenje akcijske cene do" ~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}'
       AND raw."Važenje akcijske cene od"::timestamp <= CURRENT_TIMESTAMP
       AND raw."Važenje akcijske cene do"::timestamp >= CURRENT_TIMESTAMP
      THEN (SELECT a.id FROM public."Action" a WHERE a.slug = '{PROMO_ACTION_SLUG}' LIMIT 1)
      ELSE NULL
    END AS action_id
  FROM raw
  LEFT JOIN product_assets ON product_assets.sku = raw."Šifra"
)
INSERT INTO public."Product" (
  id, sku, barcode, slug, name, description, "shortDescription",
  "colorPrimary", "colorSecondary", "groupId", "collectionId",
  "fullPrice", "salePrice", "discountPct", "actionId",
  stock, "incomingStock", "supplierStock", "deliveryDaysMin", "deliveryDaysMax",
  "allowsAssembly", "supplierId", "supplierExternalId",
  "inGoogleMerchant", "inMetaCatalog", "inTiktokCatalog", "isActive",
  "createdAt", "updatedAt"
)
SELECT
  id, sku, barcode, slug, name, description, short_description,
  color_primary, color_secondary, group_id, collection_id,
  full_price,
  sale_price,
  CASE
    WHEN sale_price IS NOT NULL AND full_price > 0
    THEN round(((full_price - sale_price) / full_price) * 100)::int
    ELSE NULL
  END,
  action_id,
  stock, 0, supplier_stock, 3, 5,
  false, supplier_id, sku,
  false, false, false, stock > 0 AND media_count > 0,
  now(), now()
FROM mapped
ON CONFLICT (sku) DO UPDATE SET
  barcode = EXCLUDED.barcode,
  slug = EXCLUDED.slug,
  name = EXCLUDED.name,
  description = EXCLUDED.description,
  "shortDescription" = EXCLUDED."shortDescription",
  "colorPrimary" = EXCLUDED."colorPrimary",
  "colorSecondary" = EXCLUDED."colorSecondary",
  "groupId" = EXCLUDED."groupId",
  "collectionId" = EXCLUDED."collectionId",
  "fullPrice" = EXCLUDED."fullPrice",
  "salePrice" = EXCLUDED."salePrice",
  "discountPct" = EXCLUDED."discountPct",
  "actionId" = EXCLUDED."actionId",
  stock = EXCLUDED.stock,
  "incomingStock" = EXCLUDED."incomingStock",
  "supplierStock" = EXCLUDED."supplierStock",
  "supplierId" = EXCLUDED."supplierId",
  "supplierExternalId" = EXCLUDED."supplierExternalId",
  "isActive" = EXCLUDED."isActive",
  "updatedAt" = now();

-- Replace only generated media rows for SKUs that have product-photo assets.
WITH media_rows AS (
  SELECT *
  FROM jsonb_to_recordset($media_json${media_json}$media_json$::jsonb) AS x(
    sku TEXT,
    media_id TEXT,
    media_order INT,
    url TEXT,
    thumb_url TEXT,
    card_url TEXT,
    pdp_url TEXT,
    alt TEXT
  )
), asset_skus AS (
  SELECT DISTINCT sku FROM media_rows
)
DELETE FROM public."ProductMedia" pm
USING public."Product" p, asset_skus s
WHERE pm."productId" = p.id
  AND p.sku = s.sku
  AND pm.id LIKE ('sa-media-' || p.sku || '-%');

WITH media_rows AS (
  SELECT *
  FROM jsonb_to_recordset($media_json${media_json}$media_json$::jsonb) AS x(
    sku TEXT,
    media_id TEXT,
    media_order INT,
    url TEXT,
    thumb_url TEXT,
    card_url TEXT,
    pdp_url TEXT,
    alt TEXT
  )
)
INSERT INTO public."ProductMedia" (id, "productId", kind, url, "thumbUrl", "cardUrl", "pdpUrl", alt, width, height, "blurDataUrl", "order")
SELECT
  media_rows.media_id,
  p.id,
  'IMAGE',
  media_rows.url,
  media_rows.thumb_url,
  media_rows.card_url,
  media_rows.pdp_url,
  media_rows.alt,
  NULL,
  NULL,
  NULL,
  media_rows.media_order
FROM media_rows
JOIN public."Product" p ON p.sku = media_rows.sku
ON CONFLICT (id) DO UPDATE SET
  "productId" = EXCLUDED."productId",
  kind = EXCLUDED.kind,
  url = EXCLUDED.url,
  "thumbUrl" = EXCLUDED."thumbUrl",
  "cardUrl" = EXCLUDED."cardUrl",
  "pdpUrl" = EXCLUDED."pdpUrl",
  alt = EXCLUDED.alt,
  "order" = EXCLUDED."order";

-- Link each product to the website category corresponding to its exact source Kategorija + Grupa.
DELETE FROM public."ProductCategory" pc
USING public."Product" p, public.svet_akcija_product_import r
WHERE pc."productId" = p.id
  AND p.sku = r."Šifra"
  AND r.import_batch = '{IMPORT_BATCH}';

WITH raw AS (
  SELECT * FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}'
), mapped AS (
  SELECT
    p.id AS product_id,
    c.id AS category_id
  FROM raw r
  JOIN public."Product" p ON p.sku = r."Šifra"
  JOIN public."Category" c ON c.path = '/' || public.svet_akcija_category_slug(r."Kategorija") || '/' || public.svet_akcija_slugify(r."Grupa")
)
INSERT INTO public."ProductCategory" ("productId", "categoryId")
SELECT product_id, category_id
FROM mapped
ON CONFLICT ("productId", "categoryId") DO NOTHING;

COMMIT;

-- Post-import review queries. Run after paste, or inspect the Results panel.
SELECT 'raw_rows' AS check_name, count(*) AS value FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}'
UNION ALL SELECT 'products_seeded', count(*) FROM public."Product" WHERE sku IN (SELECT "Šifra" FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}')
UNION ALL SELECT 'products_with_generated_media', count(DISTINCT p.sku) FROM public."ProductMedia" pm JOIN public."Product" p ON p.id = pm."productId" WHERE pm.id LIKE 'sa-media-%'
UNION ALL SELECT 'generated_media_rows', count(*) FROM public."ProductMedia" WHERE id LIKE 'sa-media-%'
UNION ALL SELECT 'categories_total', count(*) FROM public."Category"
UNION ALL SELECT 'product_category_links', count(*) FROM public."ProductCategory" pc JOIN public."Product" p ON p.id = pc."productId" WHERE p.sku IN (SELECT "Šifra" FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}')
UNION ALL SELECT 'inactive_missing_stock_or_media', count(*) FROM public."Product" p WHERE p.sku IN (SELECT "Šifra" FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}') AND p."isActive" = false
UNION ALL SELECT 'active_expired_campaign_sale_prices', count(*) FROM public."Product" p WHERE p.sku IN (SELECT "Šifra" FROM public.svet_akcija_product_import WHERE import_batch = '{IMPORT_BATCH}') AND p."isActive" = true AND p."salePrice" IS NOT NULL AND p."actionId" IS NULL;

SELECT "Bar kod", array_agg("Šifra" ORDER BY "Šifra") AS skus, count(*) AS duplicate_count
FROM public.svet_akcija_product_import
WHERE import_batch = '{IMPORT_BATCH}' AND nullif("Bar kod", '') IS NOT NULL
GROUP BY "Bar kod"
HAVING count(*) > 1;
"""


def main() -> None:
    source_wb = load_workbook(SOURCE_XLSX, data_only=False)
    product_ws = source_wb["Sheet1"]
    headers = [product_ws.cell(1, c).value for c in range(1, product_ws.max_column + 1)]

    products: list[dict[str, Any]] = []
    for row_num in range(2, product_ws.max_row + 1):
        record = {h: product_ws.cell(row_num, idx + 1).value for idx, h in enumerate(headers)}
        if any(not is_blank(record[h]) for h in headers):
            record["_source_row"] = row_num
            products.append(record)

    categories_counter = Counter(row["Kategorija"] for row in products)
    groups_counter = Counter((row["Kategorija"], row["Grupa"]) for row in products)

    by_field: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for field in ["Šifra", "Bar kod", "Kratki naziv", "Opis"]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in products:
            value = row.get(field)
            if not is_blank(value):
                grouped[str(value)].append(row)
        by_field[field] = grouped

    duplicate_sku = {k: v for k, v in by_field["Šifra"].items() if len(v) > 1}
    duplicate_barcode = {k: v for k, v in by_field["Bar kod"].items() if len(v) > 1}
    same_short_name = {k: v for k, v in by_field["Kratki naziv"].items() if len(v) > 1}
    same_description = {k: v for k, v in by_field["Opis"].items() if len(v) > 1}
    product_skus = {normalize_sku(r.get("Šifra")) for r in products if not is_blank(r.get("Šifra"))}
    product_assets_by_sku, product_asset_summary = build_product_assets(product_skus)

    data_checks: list[dict[str, Any]] = []

    def add_check(name: str, rows: list[dict[str, Any]], explanation: str, action: str) -> None:
        data_checks.append(
            {
                "Check name": name,
                "Number of affected products": len(rows),
                "Affected product codes": compact_list([r.get("Šifra") for r in rows]),
                "Affected rows": compact_list([r.get("_source_row") for r in rows]),
                "Explanation": explanation,
                "Recommended action": action,
            }
        )

    for field in REQUIRED_CHECK_FIELDS:
        rows = [r for r in products if is_blank(r.get(field))]
        add_check(
            f"Missing {field}",
            rows,
            f"Rows where source field '{field}' is empty/null.",
            "Ask client to provide values or confirm that this field should remain blank.",
        )

    add_check(
        "Duplicate Šifra",
        [r for rows in duplicate_sku.values() for r in rows],
        "Same product code appears on more than one product row.",
        "Stop import and ask client which SKU is authoritative before using Šifra as the primary key.",
    )
    add_check(
        "Duplicate Bar kod",
        [r for rows in duplicate_barcode.values() for r in rows],
        "Same barcode appears on more than one non-empty barcode row.",
        "Keep products separate by Šifra; ask client whether barcode duplication is intentional.",
    )
    missing_asset_rows = [r for r in products if normalize_sku(r.get("Šifra")) not in product_assets_by_sku]
    add_check(
        "Missing product asset folder",
        missing_asset_rows,
        "Rows where no matching folder exists in Product photos/ after trimming spaces and removing dots from folder names.",
        "Keep product without photos/full DOCX copy until assets are provided.",
    )
    data_checks.append(
        {
            "Check name": "Unmatched product asset folders",
            "Number of affected products": len(product_asset_summary["unmatchedAssetFolders"]),
            "Affected product codes": compact_list([item["normalizedSku"] for item in product_asset_summary["unmatchedAssetFolders"]]),
            "Affected rows": "",
            "Explanation": "Folders in Product photos/ whose normalized folder name does not match any catalog SKU.",
            "Recommended action": "Confirm whether these folders belong to a future catalog batch or should be renamed.",
        }
    )
    data_checks.append(
        {
            "Check name": "DOCX extraction failures",
            "Number of affected products": len(product_asset_summary["docxExtractionFailures"]),
            "Affected product codes": compact_list([item["sku"] for item in product_asset_summary["docxExtractionFailures"]]),
            "Affected rows": "",
            "Explanation": "Product folders where textutil could not extract DOCX text.",
            "Recommended action": "Fix or replace the DOCX file before publishing full product copy.",
        }
    )
    lfs_description_skus = {
        sku
        for sku, asset in product_assets_by_sku.items()
        if "broken:lfs_pointer_description" in asset.get("warnings", [])
    }
    add_check(
        "Git LFS pointer long descriptions",
        [r for r in products if normalize_sku(r.get("Šifra")) in lfs_description_skus],
        "Product DOCX files that extracted as Git LFS pointer metadata instead of real description text.",
        "Fetch the real LFS-backed DOCX files or replace them before publishing long product copy.",
    )

    empty_dc = [r for r in products if is_blank(r.get("DC (lager)"))]
    add_check(
        "Empty stock/logistics fields",
        empty_dc,
        "Rows where 'DC (lager)' is empty/null.",
        "Missing stock imports as unavailable; ask the client to provide launch-approved SKU-level availability before publishing.",
    )

    higher = []
    equal = []
    identical_or_missing_price = []
    expired_campaign = []
    invalid_campaign_dates = []
    future_campaign = []
    for r in products:
        regular = safe_number(r.get("MPC redovna"))
        sale = safe_number(r.get("Akcijska MPC"))
        if regular is not None and sale is not None:
            if sale > regular:
                higher.append(r)
            if sale == regular:
                equal.append(r)
            if sale < regular:
                status = campaign_status(
                    r.get("Važenje akcijske cene od"),
                    r.get("Važenje akcijske cene do"),
                )
                if status == "expired":
                    expired_campaign.append(r)
                elif status == "invalid":
                    invalid_campaign_dates.append(r)
                elif status == "future":
                    future_campaign.append(r)
        else:
            identical_or_missing_price.append(r)
    add_check(
        "Action price higher than regular price",
        higher,
        "Rows where numeric Akcijska MPC is greater than numeric MPC redovna.",
        "Ask client to confirm price data before publishing these SKUs.",
    )
    add_check(
        "Identical regular and action price",
        equal,
        "Rows where numeric Akcijska MPC equals numeric MPC redovna.",
        "Ask client whether these are truly action items or should display as regular-price products.",
    )
    add_check(
        "Expired campaign price window",
        expired_campaign,
        "Rows with a discounted Akcijska MPC whose validity end date is before the report generation date.",
        "Do not import these as sale prices; request current owner-approved campaign dates or publish regular price only.",
    )
    add_check(
        "Invalid campaign date window",
        invalid_campaign_dates,
        "Rows with discounted Akcijska MPC but missing or unparsable campaign validity dates.",
        "Fix campaign dates before importing a sale price.",
    )
    add_check(
        "Future campaign price window",
        future_campaign,
        "Rows with discounted Akcijska MPC whose validity start date is after the report generation date.",
        "Keep future sale prices inactive until the campaign starts.",
    )

    placeholder_rows = []
    placeholder_notes = []
    for field in headers:
        for r in products:
            value = r.get(field)
            if value is not None and str(value).strip() in PLACEHOLDERS:
                placeholder_rows.append(r)
                placeholder_notes.append(f"row {r['_source_row']} {field}={value}")
    add_check(
        "Suspicious placeholder values",
        placeholder_rows,
        "Detected values exactly matching placeholder tokens: 9, 0, /, -, N/A, NA.",
        "Do not display these values publicly unless the client confirms the value is meaningful.",
    )

    date_serial_rows = []
    for r in products:
        for field in ["Važenje akcijske cene od", "Važenje akcijske cene do"]:
            if isinstance(r.get(field), (int, float)):
                date_serial_rows.append(r)
                break
    add_check(
        "Numeric action-date values",
        date_serial_rows,
        "Rows where action validity date fields are plain numbers in the exported workbook.",
        "If present, ask whether Excel serial dates should be converted for display; keep source values unchanged.",
    )

    for field in SUSPICIOUS_FIELDS:
        counts = Counter(str(r.get(field)).strip() for r in products if not is_blank(r.get(field)))
        repeated = [(value, count) for value, count in counts.items() if count >= 10 and value not in PLACEHOLDERS]
        if repeated:
            affected = [r for r in products if str(r.get(field)).strip() in {v for v, _ in repeated}]
            data_checks.append(
                {
                    "Check name": f"Repeated non-descriptive review: {field}",
                    "Number of affected products": len(affected),
                    "Affected product codes": compact_list([r.get("Šifra") for r in affected]),
                    "Affected rows": compact_list([r.get("_source_row") for r in affected]),
                    "Explanation": "Frequently repeated values found: " + compact_list([f"{v} ({c})" for v, c in repeated], 20),
                    "Recommended action": "Confirm these are valid customer-facing filter/spec values before exposing as filters.",
                }
            )

    duplicate_checks = []

    def add_duplicate_block(label: str, grouped: dict[str, list[dict[str, Any]]], explanation: str, recommendation: str) -> None:
        for key, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
            duplicate_checks.append(
                {
                    "Check type": label,
                    "Match key": key,
                    "Count": len(rows),
                    "Product codes": compact_list([r.get("Šifra") for r in rows]),
                    "Source rows": compact_list([r.get("_source_row") for r in rows]),
                    "Explanation": explanation,
                    "Recommendation": recommendation,
                }
            )

    add_duplicate_block(
        "Duplicate Šifra rows",
        duplicate_sku,
        "Same SKU/product code appears more than once.",
        "Do not import as-is until client resolves duplicated Šifra.",
    )
    add_duplicate_block(
        "Duplicate Bar kod rows",
        duplicate_barcode,
        "Same non-empty barcode appears more than once.",
        "Use Šifra as primary key; ask client if duplicate barcode is intentional.",
    )
    add_duplicate_block(
        "Products sharing same Kratki naziv",
        same_short_name,
        "Multiple rows share the same short name.",
        "Keep as separate SKUs; review only as possible variants after client confirmation.",
    )
    add_duplicate_block(
        "Products sharing same Opis",
        same_description,
        "Multiple rows share the same short description.",
        "Keep as separate SKUs; shared description alone is not a variant rule.",
    )

    variant_suggestions = {}
    for key, rows in same_short_name.items():
        if len({str(r.get("Šifra")) for r in rows}) > 1:
            varied_fields = []
            for field in ["Opis", "Atribut 1", "Atribut 2", "Boja 1", "Boja 2", "Bar kod", "MPC redovna", "Akcijska MPC"]:
                if len({to_display(r.get(field)) for r in rows}) > 1:
                    varied_fields.append(field)
            variant_suggestions[key] = (rows, varied_fields)
    for key, (rows, varied_fields) in sorted(variant_suggestions.items(), key=lambda item: (-len(item[1][0]), item[0])):
        duplicate_checks.append(
            {
                "Check type": "Possible variant group - keep separate",
                "Match key": key,
                "Count": len(rows),
                "Product codes": compact_list([r.get("Šifra") for r in rows]),
                "Source rows": compact_list([r.get("_source_row") for r in rows]),
                "Explanation": "Same Kratki naziv with differences in: " + (", ".join(varied_fields) if varied_fields else "no reviewed fields"),
                "Recommendation": "Do not merge automatically; client must confirm variant grouping rules.",
            }
        )

    ux_plan = [
        ["Product cards", "Show image/placeholder, exact Kratki naziv, exact Opis, Akcijska MPC, optional crossed-out MPC redovna after approval, Šifra, category/group badge, and CTA.", "Keeps cards scannable while preserving key resale/e-commerce buying signals.", "Kratki naziv, Opis, Akcijska MPC, MPC redovna, Šifra, Kategorija, Grupa", "Need confirmation before showing regular price as crossed out."],
        ["Category navigation", "Use Kategorija as top-level navigation and Grupa as subcategory pages/filters.", "Matches the actual table taxonomy without inventing labels.", "Kategorija, Grupa", "Client should confirm whether these labels are final public website labels."],
        ["Filters", "Enable brand/collection, color, attributes, price range, and availability only where values are meaningful and reliable.", "Avoids noisy filters from missing or sparse fields.", "Kolekcija (brend), Boja 1, Boja 2, Atribut 1, Atribut 2, MPC redovna, Akcijska MPC, DC (lager)", "Many Boja 2 and DC (lager) values are missing; hide unreliable filters until confirmed."],
        ["Search", "Index Šifra, Kratki naziv, Opis, Kategorija, Grupa, Bar kod, and Kolekcija (brend).", "Supports customer search and internal/admin lookup.", "Šifra, Kratki naziv, Opis, Kategorija, Grupa, Bar kod, Kolekcija (brend)", "Barcode search should likely be admin/internal, not public UI copy."],
        ["Product detail page", "Use exact Kratki naziv title, exact Opis short description, price block, SKU, breadcrumb, specs table, product media gallery, and full DOCX description when available.", "Adds richer buying context without fabricating marketing copy.", "Kratki naziv, Opis, Product photos DOCX/images, Akcijska MPC, MPC redovna, Šifra, Kategorija, Grupa, Atribut 1, Atribut 2, Boja 1, Boja 2, Kolekcija (brend), Dobavljač, Bar kod", "Products without asset folders fall back to source Opis and no media."],
        ["Internal fields", "Keep Dobavljač, DC (lager), and Bar kod in admin/internal details unless client asks to show them publicly.", "These fields may confuse customers or expose operational data.", "Dobavljač, DC (lager), Bar kod", "Confirm public/private visibility per field."],
        ["Missing images", "Use Product photos media where matched by Šifra; use a neutral placeholder only for products without a matching asset folder.", "Keeps cards visual where media exists while preserving stable layout for the remaining catalog.", "Product photos folder by Šifra", "93 products currently have no matching asset folder."],
        ["Mobile layout", "Use two-column product grid where space allows, one-column on narrow screens, sticky filter drawer, and keyboard-friendly controls.", "Furniture/resale browsing needs fast scanning and accessible filtering.", "Kategorija, Grupa, filter fields, price fields", "Need frontend stack/routes before implementing UI."],
    ]

    mapping = [
        ["Šifra", "id / sku", "Stable product key and visible product code.", "Display as product code; use as primary key only if unique.", "Unique in this file."],
        ["Kategorija", "category", "Top-level catalog navigation.", "Display exact source value.", "Do not translate or rename in source dataset."],
        ["Grupa", "group / subcategory", "Subcategory navigation and filters.", "Display exact source value.", "Some group names appear under different top-level categories; keep category+group pair."],
        ["Dobavljač", "supplier", "Admin/internal supplier reference.", "Hide publicly unless client confirms.", "Useful for import/admin, not card UI."],
        ["Kolekcija (brend)", "brandOrCollection", "Brand/collection filter and spec.", "Show/filter only when meaningful and not blank.", "56 rows missing."],
        ["Opis", "shortDescription", "Product-card and detail short description.", "Display exact source value until long descriptions arrive.", "Do not invent long copy."],
        ["Kratki naziv", "title", "Product card/detail title.", "Display exact source value.", "Do not rewrite or combine with Opis."],
        ["Atribut 1", "spec.attribute1", "Specs table/filter if reliable.", "Display in specs; do not use to create variants automatically.", "Many blanks; confirm filter use."],
        ["Atribut 2", "spec.attribute2", "Specs table/filter if reliable.", "Display in specs; do not use to create variants automatically.", "Sparse field."],
        ["Boja 1", "colorPrimary", "Color spec/filter.", "Display/filter only if meaningful.", "19 rows missing."],
        ["Boja 2", "colorSecondary", "Secondary color spec/filter.", "Display/filter only if meaningful.", "171 rows missing."],
        ["DC (lager)", "stockLogistics.dcLager", "Internal stock/logistics.", "Do not show publicly unless confirmed.", "206 rows missing."],
        ["Bar kod", "barcode", "Secondary identifier/admin lookup.", "Hide from public product card; optional admin detail.", "74 rows missing and 1 duplicate barcode group."],
        ["MPC redovna", "regularPrice", "Regular/old price.", "Display as old/crossed-out price only after business confirmation.", "Preserved exactly; no discount calculation added."],
        ["Akcijska MPC", "salePrice", "Primary displayed price.", "Use as primary displayed price when present.", "All rows have a value."],
        ["Važenje akcijske cene od", "saleValidFrom", "Promo validity start.", "Keep source date; display only if client approves.", "Exported XLSX stores date cells as serial-backed Excel dates."],
        ["Važenje akcijske cene do", "saleValidTo", "Promo validity end.", "Keep source date; display only if client approves.", "Exported XLSX stores date cells as serial-backed Excel dates."],
    ]

    product_json = []
    duplicate_barcode_values = set(duplicate_barcode.keys())
    for r in products:
        sku = normalize_sku(r.get("Šifra"))
        asset = product_assets_by_sku.get(sku)
        long_description = asset["longDescription"] if asset and asset.get("longDescription") else to_display(r.get("Opis"))
        description_source = "product_assets_docx" if asset and asset.get("longDescription") else "source_opis"
        media = asset["media"] if asset else []
        flags = []
        for field in REQUIRED_CHECK_FIELDS:
            if is_blank(r.get(field)):
                flags.append(f"missing:{field}")
        if asset is None:
            flags.append("missing:product_assets")
        elif not media:
            flags.append("missing:product_images")
        if asset and not asset.get("longDescription"):
            flags.append("missing:product_long_description")
        if asset:
            flags.extend(asset["warnings"])
        if not is_blank(r.get("Bar kod")) and str(r.get("Bar kod")) in duplicate_barcode_values:
            flags.append("duplicate:Bar kod")
        regular = safe_number(r.get("MPC redovna"))
        sale = safe_number(r.get("Akcijska MPC"))
        if regular is not None and sale is not None:
            if sale > regular:
                flags.append("price:action_higher_than_regular")
            if sale == regular:
                flags.append("price:regular_equals_action")
            if sale < regular:
                status = campaign_status(
                    r.get("Važenje akcijske cene od"),
                    r.get("Važenje akcijske cene do"),
                )
                if status != "current":
                    flags.append(f"campaign:{status}")
        for field in headers:
            value = r.get(field)
            if value is not None and str(value).strip() in PLACEHOLDERS:
                flags.append(f"suspicious_placeholder:{field}")
        product_json.append(
            {
                "source": as_source_record(r, headers),
                "website_mapping": {
                    "id": to_display(r.get("Šifra")),
                    "title": to_display(r.get("Kratki naziv")),
                    "shortDescription": to_display(r.get("Opis")),
                    "longDescription": long_description,
                    "descriptionSource": description_source,
                    "category": to_display(r.get("Kategorija")),
                    "group": to_display(r.get("Grupa")),
                    "regularPrice": to_display(r.get("MPC redovna")),
                    "salePrice": to_display(r.get("Akcijska MPC")),
                    "brandOrCollection": to_display(r.get("Kolekcija (brend)")),
                    "colorPrimary": to_display(r.get("Boja 1")),
                    "colorSecondary": to_display(r.get("Boja 2")),
                    "barcode": to_display(r.get("Bar kod")),
                    "sku": to_display(r.get("Šifra")),
                    "media": media,
                },
                "flags": flags,
            }
        )

    categories_json = [
        {
            "Kategorija": category,
            "productCount": count,
            "percentageOfCatalog": round(count / len(products), 4),
        }
        for category, count in sorted(categories_counter.items(), key=lambda item: item[0])
    ]
    groups_json = [
        {
            "Kategorija": category,
            "Grupa": group,
            "productCount": count,
            "percentageOfCatalog": round(count / len(products), 4),
        }
        for (category, group), count in sorted(groups_counter.items(), key=lambda item: (item[0][0], item[0][1]))
    ]
    product_assets_json = {
        "summary": product_asset_summary,
        "assets": [product_assets_by_sku[sku] for sku in sorted(product_assets_by_sku, key=natural_key)],
    }
    manifest_entries = [
        {
            "sku": sku,
            "mediaId": media["id"],
            "mediaOrder": media["order"],
            "localSourcePath": media["absoluteLocalPath"],
            "relativeSourcePath": media["localPath"],
            "storagePath": media["storagePath"],
        }
        for sku in sorted(product_assets_by_sku, key=natural_key)
        for media in product_assets_by_sku[sku]["media"]
    ]
    media_upload_manifest_json = {
        "summary": {
            "productCountWithMedia": len({entry["sku"] for entry in manifest_entries}),
            "mediaFileCount": len(manifest_entries),
            "storagePathShape": "products/{sku}/{order}-{slugified-original-name}.{ext}",
        },
        "entries": manifest_entries,
    }

    json_dump(PRODUCTS_JSON_OUT, product_json)
    json_dump(CATEGORIES_JSON_OUT, categories_json)
    json_dump(GROUPS_JSON_OUT, groups_json)
    json_dump(CHECKS_JSON_OUT, data_checks)
    json_dump(PRODUCT_ASSETS_JSON_OUT, product_assets_json)
    json_dump(MEDIA_UPLOAD_MANIFEST_JSON_OUT, media_upload_manifest_json)
    SUPABASE_SQL_OUT.write_text(
        render_supabase_import_sql(products, headers, product_assets_by_sku, duplicate_barcode_values),
        encoding="utf-8",
    )

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    products_sheet = wb.create_sheet("Products_exact")
    products_sheet.append(headers)
    for source_row_idx in range(2, product_ws.max_row + 1):
        row = [product_ws.cell(source_row_idx, col_idx).value for col_idx in range(1, product_ws.max_column + 1)]
        if any(not is_blank(v) for v in row):
            products_sheet.append(row)
            target_row_idx = products_sheet.max_row
            for col_idx in range(1, product_ws.max_column + 1):
                source_cell = product_ws.cell(source_row_idx, col_idx)
                target_cell = products_sheet.cell(target_row_idx, col_idx)
                target_cell.number_format = source_cell.number_format
    for cell in products_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="173E43")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    products_sheet.freeze_panes = "A2"
    products_sheet.auto_filter.ref = products_sheet.dimensions
    for col_idx, header in enumerate(headers, 1):
        width = 14 if header not in {"Opis", "Dobavljač"} else 32
        products_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    for row in products_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_sheet(
        wb,
        "Categories",
        ["Kategorija / category", "Product count", "Percentage of catalog"],
        [[item["Kategorija"], item["productCount"], item["percentageOfCatalog"]] for item in categories_json],
    )
    add_sheet(
        wb,
        "Groups",
        ["Kategorija / category", "Grupa / group/subcategory", "Product count", "Percentage of catalog"],
        [[item["Kategorija"], item["Grupa"], item["productCount"], item["percentageOfCatalog"]] for item in groups_json],
    )
    add_sheet(
        wb,
        "Data_checks",
        ["Check name", "Number of affected products", "Affected product codes", "Affected rows", "Explanation", "Recommended action"],
        [[c["Check name"], c["Number of affected products"], c["Affected product codes"], c["Affected rows"], c["Explanation"], c["Recommended action"]] for c in data_checks],
        {1: 34, 2: 18, 3: 52, 4: 32, 5: 54, 6: 54},
    )
    add_sheet(
        wb,
        "Duplicate_checks",
        ["Check type", "Match key", "Count", "Product codes", "Source rows", "Explanation", "Recommendation"],
        [[c["Check type"], c["Match key"], c["Count"], c["Product codes"], c["Source rows"], c["Explanation"], c["Recommendation"]] for c in duplicate_checks],
        {1: 34, 2: 36, 3: 10, 4: 50, 5: 34, 6: 52, 7: 52},
    )
    add_sheet(
        wb,
        "UX_plan",
        ["Area", "Recommendation", "Reason", "Source fields used", "Risk / question for client"],
        ux_plan,
        {1: 22, 2: 70, 3: 48, 4: 48, 5: 48},
    )
    add_sheet(
        wb,
        "Implementation_mapping",
        ["Source field", "Website field", "Usage", "Display rule", "Notes"],
        mapping,
        {1: 28, 2: 26, 3: 48, 4: 52, 5: 52},
    )
    add_sheet(
        wb,
        "Product_assets",
        ["SKU", "Source folder", "DOCX file", "Description chars", "Image count", "Warnings"],
        [
            [
                asset["sku"],
                asset["sourceFolder"],
                asset["docxFile"],
                asset["descriptionCharCount"],
                len(asset["media"]),
                ", ".join(asset["warnings"]),
            ]
            for asset in product_assets_json["assets"]
        ],
        {1: 14, 2: 26, 3: 44, 4: 18, 5: 14, 6: 48},
    )
    add_sheet(
        wb,
        "Media_upload_manifest",
        ["SKU", "Media ID", "Order", "Relative source path", "Storage path"],
        [
            [
                entry["sku"],
                entry["mediaId"],
                entry["mediaOrder"],
                entry["relativeSourcePath"],
                entry["storagePath"],
            ]
            for entry in manifest_entries
        ],
        {1: 14, 2: 26, 3: 10, 4: 60, 5: 60},
    )

    wb.save(WORKBOOK_OUT)

    print(json.dumps({
        "workbook": str(WORKBOOK_OUT),
        "productsJson": str(PRODUCTS_JSON_OUT),
        "categoriesJson": str(CATEGORIES_JSON_OUT),
        "groupsJson": str(GROUPS_JSON_OUT),
        "checksJson": str(CHECKS_JSON_OUT),
        "productAssetsJson": str(PRODUCT_ASSETS_JSON_OUT),
        "mediaUploadManifestJson": str(MEDIA_UPLOAD_MANIFEST_JSON_OUT),
        "supabaseSql": str(SUPABASE_SQL_OUT),
        "products": len(products),
        "columns": len(headers),
        "categories": len(categories_counter),
        "groups": len(groups_counter),
        "matchedAssetSkus": product_asset_summary["matchedAssetSkus"],
        "missingAssetSkus": len(product_asset_summary["missingAssetSkus"]),
        "mediaFiles": len(manifest_entries),
        "dataChecks": len(data_checks),
        "duplicateChecks": len(duplicate_checks),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
